from tkinter import * 
import sqlite3
import tkinter.messagebox
import serial
from datetime import datetime
import os
import openpyxl
def main_app():
    current_directory=os.getcwd()
    date = datetime.today().strftime('%d-%m-%Y')
    port=str(port_number.get()).upper()
    ser = serial.Serial(port, baudrate=9600, timeout=1)

    time = datetime.today().strftime('%H:%M')

    file_name=str(date)+'.xlsx'
    if not os.path.isfile(file_name):#create excel file if not exsist
        wb = openpyxl.Workbook()
        wb.save(os.path.join(current_directory, file_name))

    workbook = openpyxl.load_workbook(file_name)#open excel file
    connect = sqlite3.connect('RFID.db')#open database
    cursor = connect.cursor()
    cursor.execute(
            'SELECT Employe_ID,Employe_NAME from Employe')#select all employe names and id's
    myself = cursor.fetchall()
    connect.commit()
    connect.close()
    
    pointer=workbook.active
    title1=pointer.cell(row=1,column=1)#defining column name
    title1.value="Employe_ID"

    title2=pointer.cell(row=1,column=2)#defining column name
    title2.value="Employe_NAME"

    title3=pointer.cell(row=1,column=3)#defining column name
    title3.value="Time"
    for i, row in enumerate(myself): # inserting id and employ names
        for j, value in enumerate(row):
            title1=pointer.cell(row=i+2,column=j+1)#column name
            title1.value=str(row[j])
            #worksheet.write(i+1, j, str(row[j]))
    workbook.save(file_name)
################################################################################
    win.destroy() #destroy previous window

    def rfid():#rfid tag detector code
        while 1:
            port_data = ser.readline().decode("ascii")
            if(port_data):
                rfid_value = port_data[10:21]
                return rfid_value
    try:
        while True:#updation in excel file
            #win.destroy(
            rfid_value=rfid()
            print(rfid_value)
            connect = sqlite3.connect('RFID.db')
            cursor = connect.cursor()
            cursor.execute('SELECT Employe_ID  from Employe where RFID_tag =?', (rfid_value,))
            data=cursor.fetchall()
            employ_id=data[0][0]
            wb1 =openpyxl.load_workbook(file_name)
            pointer=wb1.active
            length_row=pointer.max_row
            for i in range(2,length_row+1):
                id_=pointer.cell(row = i, column = 1)
                id_=id_.value
                if int(id_)==int(employ_id):
                    updated_time=pointer.cell(row=i,column=3)
                    updated_time.value=time
                    wb1.save(file_name)
    except:
        tkinter.messagebox.showinfo('Connection broken','devise is not connected properly')


win = Tk()
win.geometry('400x130')
port_number= StringVar()
label = Label(win, text='PORT NUMBER:',
                  font='Times 16 bold', fg='black')
label.place(x=40, y=10)
port_info = Entry(win, justify=RIGHT, width=30,
                                textvariable=port_number, bg='grey39', fg='white', font='Times 16 bold')
port_info.place(x=40, y=50)
enter_button=Button(win, text='ENTER', bg='grey15', fg='white',
                               font='Times 16 bold', command=main_app, height=1, width=8, padx=3, pady=3)
enter_button.place(x=280,y=80)
    
win.mainloop()
